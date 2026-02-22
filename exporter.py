import pandas as pd
from fpdf import FPDF
import os
import matplotlib
import matplotlib.pyplot as plt
import tempfile

# Use non-interactive backend for server-side generation
matplotlib.use('Agg')

class RCAExporter:
    def __init__(self, output_dir=None):
        if output_dir is None:
            output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports')
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def _generate_visuals(self, results):
        """Generates matplotlib charts and returns paths to temporary images."""
        temp_files = []
        
        # 1. Timeline Chart
        ts = results.get('time_series', {})
        if ts and ts.get('labels'):
            plt.figure(figsize=(10, 5))
            plt.style.use('dark_background')
            plt.plot(ts['labels'], ts['counts'], color='#3b82f6', linewidth=2.5, marker='o', markersize=4)
            plt.fill_between(ts['labels'], ts['counts'], color='#3b82f6', alpha=0.15)
            
            plt.title("Incident Frequency Timeline", color='#f8fafc', pad=20, fontsize=14, fontweight='bold')
            plt.xlabel("Time Window", color='#94a3b8', fontsize=10)
            plt.ylabel("Incident Count", color='#94a3b8', fontsize=10)
            
            # Smart tick spacing
            if len(ts['labels']) > 15:
                plt.xticks(range(0, len(ts['labels']), len(ts['labels'])//10), rotation=45, ha='right', fontsize=9)
            else:
                plt.xticks(rotation=45, ha='right', fontsize=9)
                
            plt.grid(True, linestyle='--', alpha=0.15)
            plt.tight_layout()
            
            tf = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            plt.savefig(tf.name, transparent=False, facecolor='#0f172a') # Match dark theme
            plt.close()
            temp_files.append(('timeline', tf.name))

        # 2. Service Matrix Chart
        matrix = results.get('service_matrix', {})
        if matrix:
            plt.figure(figsize=(10, 6))
            plt.style.use('dark_background')
            services = list(matrix.keys())
            # Find all possible status keys
            all_statuses = set()
            for s in matrix.values():
                all_statuses.update(s.keys())
            
            all_statuses = sorted(list(all_statuses))
            bottom = [0] * len(services)
            colors = ['#ef4444', '#10b981', '#3b82f6', '#f59e0b', '#6366f1', '#ec4899', '#8b5cf6']
            
            for i, status in enumerate(all_statuses):
                vals = [matrix[s].get(status, 0) for s in services]
                plt.bar(services, vals, bottom=bottom, label=status, color=colors[i % len(colors)])
                bottom = [b + v for b, v in zip(bottom, vals)]
            
            plt.title("Service Reliability Distribution", color='#f8fafc', pad=20, fontsize=14, fontweight='bold')
            plt.xlabel("Services / Categories", color='#94a3b8', fontsize=10)
            plt.ylabel("Event Count", color='#94a3b8', fontsize=10)
            
            plt.xticks(rotation=45, ha='right', fontsize=9)
            plt.legend(frameon=True, loc='upper right', fontsize=9, facecolor='#1e293b')
            plt.tight_layout()
            
            tf = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            plt.savefig(tf.name, transparent=False, facecolor='#0f172a')
            plt.close()
            temp_files.append(('matrix', tf.name))

        return temp_files

    def to_excel(self, results, report_id):
        filepath = os.path.join(self.output_dir, f"{report_id}.xlsx")
        # Flattened Summary
        summary_data = {
            "Metric": ["Report ID", "Date", "Total Records", "Confidence", "AI Narrative", "Root Cause"],
            "Value": [
                results.get('report_id'), 
                results.get('date'), 
                results.get('total_records'), 
                f"{results.get('confidence_score')}%", 
                results.get('summary', 'Analysis generated successfully.'),
                results.get('root_cause_prediction')
            ]
        }
        
        temp_plots = self._generate_visuals(results)
        
        try:
            from openpyxl import Workbook
            from openpyxl.drawing.image import Image
            
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Executive Summary', index=False)
            
            matrix = results.get('service_matrix', {})
            if matrix:
                matrix_df = pd.DataFrame(matrix).T.fillna(0)
                matrix_df.to_excel(writer, sheet_name='Service Distribution')
                
            anom_list = results.get('anomalies', [])
            anom_df = pd.DataFrame({"Anomalies Detected": anom_list if anom_list else ["No anomalies detected"]})
            anom_df.to_excel(writer, sheet_name='Anomalies', index=False)
            
            reco_list = results.get('recommendations', [])
            reco_df = pd.DataFrame({"AI Recommendations": reco_list if reco_list else ["Continue routine health monitoring"]})
            reco_df.to_excel(writer, sheet_name='Recommendations', index=False)
            
            # --- Visual Insights Sheet ---
            if temp_plots:
                ws = writer.book.create_sheet("Visual Insights")
                ws.cell(row=1, column=1, value="AI Diagnostic Patterns")
                
                curr_row = 2
                for type, path in temp_plots:
                    img = Image(path)
                    # Scale image for Excel
                    img.width = 700
                    img.height = 350
                    ws.add_image(img, f"B{curr_row}")
                    curr_row += 20 # Space for next image
            
            writer.close()
            return filepath
        finally:
            # Cleanup temp files
            for _, path in temp_plots:
                if os.path.exists(path):
                    os.remove(path)

    def to_pdf(self, results, report_id):
        filepath = os.path.join(self.output_dir, f"{report_id}.pdf")
        temp_plots = []
        try:
            temp_plots = self._generate_visuals(results)
            
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=20)
            pdf.add_page()
            
            # Header
            pdf.set_fill_color(30, 41, 59) # Dark Navy
            pdf.rect(0, 0, 210, 45, 'F')
            
            pdf.set_font("Helvetica", 'B', 22)
            pdf.set_text_color(248, 250, 252)
            pdf.ln(5)
            pdf.cell(0, 15, "AUTONOMOUS RCA REPORT", ln=True, align='C')
            
            pdf.set_font("Helvetica", 'B', 10)
            pdf.set_text_color(148, 163, 184)
            pdf.cell(0, 5, f"SYSTEM GENERATED | ID: {str(report_id)}", ln=True, align='C')
            pdf.cell(0, 5, f"DATE: {str(results.get('date'))}", ln=True, align='C')
            pdf.ln(15)
            
            # 1. Executive Summary
            pdf.set_font("Helvetica", 'B', 14)
            pdf.set_text_color(15, 23, 42)
            pdf.cell(0, 10, "1. EXECUTIVE SUMMARY", ln=True)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(5)
            
            pdf.set_font("Helvetica", '', 11)
            pdf.set_text_color(30, 41, 59)
            pdf.cell(0, 8, f"Analysis Scope: {results.get('total_records')} log entries processed.", ln=True)
            pdf.cell(0, 8, f"AI Diagnostic Confidence: {results.get('confidence_score')}%", ln=True)
            pdf.set_font("Helvetica", 'I', 11)
            pdf.multi_cell(190, 8, f"AI Narrative: {results.get('summary', 'Deep scan completed.')}")
            pdf.ln(5)
            
            # 2. Diagnostic Patterns (Visuals)
            if temp_plots:
                pdf.add_page()
                pdf.set_font("Helvetica", 'B', 14)
                pdf.cell(0, 10, "2. DIAGNOSTIC PATTERNS", ln=True)
                pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                pdf.ln(10)
                
                for type, path in temp_plots:
                    pdf.image(path, w=180)
                    pdf.ln(10)
            
            # 3. Root Cause Conclusion
            pdf.add_page()
            pdf.set_font("Helvetica", 'B', 14)
            pdf.cell(0, 10, "3. ROOT CAUSE CONCLUSION", ln=True)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(5)
            pdf.set_font("Helvetica", 'I', 11)
            pdf.set_fill_color(241, 245, 249)
            prediction = str(results.get('root_cause_prediction', 'N/A'))
            pdf.multi_cell(190, 8, prediction, fill=True)
            pdf.ln(5)
            
            # 4. Key Observations (Anomalies)
            pdf.set_font("Helvetica", 'B', 14)
            pdf.cell(0, 10, "4. SYSTEM ANOMALIES", ln=True)
            pdf.set_font("Helvetica", '', 11)
            anom_list = results.get('anomalies', [])
            if not anom_list:
                pdf.cell(0, 8, "- No anomalies detected.", ln=True)
            for anomaly in anom_list:
                pdf.multi_cell(190, 8, f"[*] {str(anomaly)}")
            pdf.ln(5)
            
            # 5. Action Items (Recommendations)
            pdf.set_font("Helvetica", 'B', 14)
            pdf.set_text_color(21, 128, 61)
            pdf.cell(0, 10, "5. PROACTIVE RECOMMENDATIONS", ln=True)
            pdf.set_font("Helvetica", 'B', 11)
            reco_list = results.get('recommendations', [])
            if not reco_list:
                pdf.cell(0, 8, "-> Continue routine health monitoring.", ln=True)
            for reco in reco_list:
                pdf.multi_cell(190, 8, f"[ACTION] -> {str(reco)}")
            
            # Footer
            pdf.set_y(-20)
            pdf.set_font("Helvetica", 'I', 8)
            pdf.set_text_color(148, 163, 184)
            pdf.cell(0, 10, "Confidential - AI Operations Center Internal Report", align='C')
            
            pdf.output(filepath)
            return filepath
        except Exception as e:
            print(f"PDF Export Error: {str(e)}")
            # Fallback
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, "RCA Report (Emergency Fallback)", ln=True)
            pdf.ln(10)
            pdf.set_font("Arial", '', 11)
            pdf.multi_cell(0, 8, f"Summary: {str(results.get('summary'))}")
            pdf.ln(5)
            pdf.multi_cell(0, 8, f"Root Cause: {str(results.get('root_cause_prediction'))}")
            pdf.output(filepath)
            return filepath
        finally:
            # Cleanup temp files
            for _, path in temp_plots:
                if os.path.exists(path):
                    os.remove(path)
